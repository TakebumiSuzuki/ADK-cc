"""Build a native chart from scratch and embed it in a slide (raw-XML, no python-pptx).

Why this exists: set_chart_data.py *reuses* a chart that already lives in the
template — it forks the sample chart part and rewrites its data, inheriting the
sample's plot type, axes and colours. That needs the template to ship a chart
sample of the matching plot type. When the template has NO chart sample at all
(or none of the needed plot type), there is nothing to fork. This script is that
fallback: it authors a complete, theme-consistent chart part (chart{N}.xml + its
embedded .xlsx + rels + Content_Types) from the YAML data and inserts a
<p:graphicFrame> into the slide that references it.

Two situations call for it (see [Charts] in SKILL.md):
  1. The template ships NO chart sample at all — nothing to copy.
  2. The template has a chart, but of the WRONG plot type (e.g. only a bar sample
     but the YAML wants a pie). Here you still want the deck's *placement* — so map
     the slide to the closest chart sample (copy it, chart and all), then run this
     script: it **takes over the existing chart's frame in place** — inheriting its
     exact position and size — removes the old (wrong-type) chart, and drops the
     new from-scratch chart into that same slot. Routing rule: matching type ->
     set_chart_data.py (reuse); type mismatch or no chart -> add_chart.py.

In-place takeover is automatic: if the slide already holds a chart graphicFrame,
this script reuses its geometry and replaces it (the old chart part is orphaned for
clean.py to drop). If the slide has no chart, the frame is placed at --area-in, or
a default content box below a title.

Theme consistency without reading the theme: series are coloured with
`<a:schemeClr val="accentN"/>` references, so the chart picks up the deck's real
accent palette automatically (and follows it if the theme is swapped) — no colour
is hard-coded and theme1.xml is never parsed.

Prefer set_chart_data.py whenever a chart sample of the MATCHING type exists (it
inherits the template's exact axis/label styling). Use this when no sample exists,
or to swap a sample chart to a different plot type while keeping its placement.

Usage:
    # Preferred: read the data entry straight from the slide-deck YAML (no temp JSON).
    python add_chart.py <unpacked_dir> <slideN.xml> --yaml <deck.yaml> --index N [--data KEY]
        [--area-in "x,y,w,h"] [--title "Chart title"]
    # Or pass a pre-extracted entry as JSON.
    python add_chart.py <unpacked_dir> <slideN.xml> --data-json <entry.json>
        [--area-in "x,y,w,h"]   # frame box in inches; overrides inherited/default geometry
        [--title "Chart title"] # optional chart title; default: none (slide title carries it)

With `--yaml/--index`, the script extracts `slides[N].data[KEY]` itself (KEY may
be omitted when the slide has a single data entry), so values are never retyped
by hand. `entry.json` is one YAML `data:` entry, e.g.
    {"type":"bar_chart","categories":["Q1","Q2"],
     "series":[{"name":"2024","values":[10,20]},{"name":"2025","values":[12,25]}]}
    {"type":"line_chart","categories":["Jan","Feb","Mar"],
     "series":[{"name":"Visits","values":[5,8,6]}]}
    {"type":"pie_chart","categories":["Ent","SMB"],
     "series":[{"name":"Share","values":[70,30]}]}
    {"type":"histogram","bins":["0-20","20-40"],"frequencies":[3,12]}

Supported types: bar_chart, line_chart, pie_chart, histogram.
"""

import argparse
import json
import re
from pathlib import Path

import defusedxml.minidom as minidom

CHART_URI = "http://schemas.openxmlformats.org/drawingml/2006/chart"
from xml.sax.saxutils import escape

EMU_PER_INCH = 914400
CAT_AX_ID = "111111111"
VAL_AX_ID = "222222222"


# --------------------------------------------------------------------------- #
# Normalise the YAML entry into (kind, categories, [(series_name, [values])])
# --------------------------------------------------------------------------- #
def _normalize(entry):
    etype = entry.get("type")
    if etype == "histogram":
        kind = "histogram"
        cats = [str(b) for b in entry.get("bins", [])]
        series = [(entry.get("y_axis", "Frequency"),
                   [float(v) for v in entry.get("frequencies", [])])]
    elif etype == "line_chart":
        kind = "line"
        cats = [str(c) for c in entry.get("categories", [])]
        series = [(s.get("name", f"Series {i+1}"),
                   [float(v) for v in s.get("values", [])])
                  for i, s in enumerate(entry.get("series", []))]
    elif etype == "bar_chart":
        kind = "bar"
        cats = [str(c) for c in entry.get("categories", [])]
        series = [(s.get("name", f"Series {i+1}"),
                   [float(v) for v in s.get("values", [])])
                  for i, s in enumerate(entry.get("series", []))]
    elif etype == "pie_chart":
        kind = "pie"
        cats = [str(c) for c in entry.get("categories", [])]
        ser_list = entry.get("series", [])
        if not ser_list:
            raise SystemExit("Error: pie_chart entry has no series")
        # A pie shows one series (parts of a whole); use the first if more are given.
        s0 = ser_list[0]
        series = [(s0.get("name", "Series 1"),
                   [float(v) for v in s0.get("values", [])])]
    else:
        raise SystemExit(f"Error: add_chart.py cannot build chart type {etype!r} "
                         "(supported: bar_chart, line_chart, pie_chart, histogram)")
    if not series:
        raise SystemExit("Error: chart entry has no series/frequencies")
    if not cats:
        raise SystemExit("Error: chart entry has no categories/bins")
    return kind, cats, series


def _num(v):
    f = float(v)
    return str(int(f)) if f.is_integer() else repr(f)


def _col_letter(i):  # 0->B, 1->C ... (data columns start at B; A is categories)
    return chr(ord("B") + i)


def _next_index(dir_path, stem, ext):
    if not dir_path.exists():
        return 1
    nums = [int(m.group(1)) for f in dir_path.glob(f"{stem}*{ext}")
            if (m := re.match(rf"{stem}(\d+){re.escape(ext)}", f.name))]
    return max(nums, default=0) + 1


# --------------------------------------------------------------------------- #
# Chart XML authoring
# --------------------------------------------------------------------------- #
def _str_cache(formula, values):
    pts = "".join(
        f'<c:pt idx="{i}"><c:v>{escape(str(v))}</c:v></c:pt>'
        for i, v in enumerate(values)
    )
    return (f'<c:strRef><c:f>{escape(formula)}</c:f>'
            f'<c:strCache><c:ptCount val="{len(values)}"/>{pts}</c:strCache></c:strRef>')


def _num_cache(formula, values):
    pts = "".join(
        f'<c:pt idx="{i}"><c:v>{_num(v)}</c:v></c:pt>'
        for i, v in enumerate(values)
    )
    return (f'<c:numRef><c:f>{escape(formula)}</c:f>'
            f'<c:numCache><c:formatCode>General</c:formatCode>'
            f'<c:ptCount val="{len(values)}"/>{pts}</c:numCache></c:numRef>')


def _series_xml(kind, idx, name, cats, vals):
    col = _col_letter(idx)
    accent = f"accent{(idx % 6) + 1}"
    n = len(cats)
    tx = f'<c:tx>{_str_cache(f"Sheet1!${col}$1", [name])}</c:tx>'
    cat = f'<c:cat>{_str_cache(f"Sheet1!$A$2:$A${n + 1}", cats)}</c:cat>'
    val = f'<c:val>{_num_cache(f"Sheet1!${col}$2:${col}${n + 1}", vals)}</c:val>'

    if kind == "pie":
        # Pie slices are coloured by varyColors at the plot level (one per category),
        # so the single series carries no per-series fill.
        return (f'<c:ser><c:idx val="{idx}"/><c:order val="{idx}"/>{tx}'
                f'{cat}{val}</c:ser>')

    if kind == "line":
        sppr = (f'<c:spPr><a:ln w="28575" cap="rnd">'
                f'<a:solidFill><a:schemeClr val="{accent}"/></a:solidFill>'
                f'<a:round/></a:ln><a:effectLst/></c:spPr>')
        marker = '<c:marker><c:symbol val="none"/></c:marker>'
        return (f'<c:ser><c:idx val="{idx}"/><c:order val="{idx}"/>{tx}'
                f'{sppr}{marker}{cat}{val}<c:smooth val="0"/></c:ser>')

    # bar / histogram
    sppr = (f'<c:spPr><a:solidFill><a:schemeClr val="{accent}"/></a:solidFill>'
            f'<a:ln><a:noFill/></a:ln></c:spPr>')
    return (f'<c:ser><c:idx val="{idx}"/><c:order val="{idx}"/>{tx}'
            f'{sppr}{cat}{val}</c:ser>')


def _plot_block(kind, sers_xml, n_series):
    if kind == "pie":
        # No axes; varyColors gives each slice its own theme colour automatically.
        return (f'<c:pieChart><c:varyColors val="1"/>{sers_xml}'
                f'<c:firstSliceAng val="0"/></c:pieChart>')

    if kind == "line":
        return (f'<c:lineChart><c:grouping val="standard"/><c:varyColors val="0"/>'
                f'{sers_xml}<c:marker val="1"/>'
                f'<c:axId val="{CAT_AX_ID}"/><c:axId val="{VAL_AX_ID}"/></c:lineChart>')

    gap = "30" if kind == "histogram" else "150"
    overlap = '<c:overlap val="-27"/>' if (kind == "bar" and n_series > 1) else ""
    return (f'<c:barChart><c:barDir val="col"/><c:grouping val="clustered"/>'
            f'<c:varyColors val="0"/>{sers_xml}'
            f'<c:gapWidth val="{gap}"/>{overlap}'
            f'<c:axId val="{CAT_AX_ID}"/><c:axId val="{VAL_AX_ID}"/></c:barChart>')


def _axes_block():
    cat_ax = (f'<c:catAx><c:axId val="{CAT_AX_ID}"/>'
              f'<c:scaling><c:orientation val="minMax"/></c:scaling>'
              f'<c:delete val="0"/><c:axPos val="b"/>'
              f'<c:majorTickMark val="out"/><c:minorTickMark val="none"/>'
              f'<c:tickLblPos val="nextTo"/><c:crossAx val="{VAL_AX_ID}"/>'
              f'<c:crosses val="autoZero"/><c:auto val="1"/>'
              f'<c:lblAlgn val="ctr"/><c:lblOffset val="100"/>'
              f'<c:noMultiLvlLbl val="0"/></c:catAx>')
    val_ax = (f'<c:valAx><c:axId val="{VAL_AX_ID}"/>'
              f'<c:scaling><c:orientation val="minMax"/></c:scaling>'
              f'<c:delete val="0"/><c:axPos val="l"/><c:majorGridlines/>'
              f'<c:majorTickMark val="out"/><c:minorTickMark val="none"/>'
              f'<c:tickLblPos val="nextTo"/><c:crossAx val="{CAT_AX_ID}"/>'
              f'<c:crosses val="autoZero"/></c:valAx>')
    return cat_ax + val_ax


def build_chart_xml(kind, cats, series, title, rid_external):
    sers_xml = "".join(
        _series_xml(kind, i, name, cats, vals)
        for i, (name, vals) in enumerate(series)
    )
    plot = _plot_block(kind, sers_xml, len(series))
    axes = "" if kind == "pie" else _axes_block()

    if title:
        title_xml = (f'<c:title><c:tx><c:rich><a:bodyPr/><a:lstStyle/>'
                     f'<a:p><a:r><a:t>{escape(title)}</a:t></a:r></a:p></c:rich></c:tx>'
                     f'<c:overlay val="0"/></c:title><c:autoTitleDeleted val="0"/>')
    else:
        title_xml = '<c:autoTitleDeleted val="1"/>'

    legend = ('<c:legend><c:legendPos val="b"/><c:overlay val="0"/></c:legend>'
              if (len(series) > 1 or kind == "pie") else "")

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<c:chartSpace '
        'xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<c:date1904 val="0"/>'
        '<c:chart>'
        f'{title_xml}'
        f'<c:plotArea><c:layout/>{plot}{axes}</c:plotArea>'
        f'{legend}'
        '<c:plotVisOnly val="1"/><c:dispBlanksAs val="gap"/>'
        '</c:chart>'
        '<c:txPr><a:bodyPr/><a:lstStyle/>'
        '<a:p><a:pPr><a:defRPr sz="1400"/></a:pPr><a:endParaRPr lang="en-US"/></a:p>'
        '</c:txPr>'
        f'<c:externalData r:id="{rid_external}"><c:autoUpdate val="0"/></c:externalData>'
        '</c:chartSpace>'
    )


# --------------------------------------------------------------------------- #
# Embedded workbook (so PowerPoint's "Edit Data" matches the cached values)
# --------------------------------------------------------------------------- #
def write_workbook(target, cats, series):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=None)
    for si, (sname, _vals) in enumerate(series):
        ws.cell(row=1, column=2 + si, value=sname)
    for ri, cat in enumerate(cats):
        ws.cell(row=2 + ri, column=1, value=cat)
        for si, (_sname, vals) in enumerate(series):
            ws.cell(row=2 + ri, column=2 + si,
                    value=(vals[ri] if ri < len(vals) else None))
    wb.save(str(target))


# --------------------------------------------------------------------------- #
# Package plumbing (Content_Types, slide rels, slide spTree) — string-edited,
# mirroring add_slide.py's style.
# --------------------------------------------------------------------------- #
def _ensure_content_types(unpacked, chart_part):
    ct_path = unpacked / "[Content_Types].xml"
    ct = ct_path.read_text(encoding="utf-8")
    changed = False

    if 'Extension="xlsx"' not in ct:
        default = ('<Default Extension="xlsx" '
                   'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"/>')
        ct = ct.replace("</Types>", f"  {default}\n</Types>")
        changed = True

    if chart_part not in ct:
        override = (f'<Override PartName="{chart_part}" '
                    'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
        ct = ct.replace("</Types>", f"  {override}\n</Types>")
        changed = True

    if changed:
        ct_path.write_text(ct, encoding="utf-8")


def _add_slide_chart_rel(unpacked, slide_name, chart_file):
    rels_dir = unpacked / "ppt" / "slides" / "_rels"
    rels_dir.mkdir(exist_ok=True)
    rels_path = rels_dir / f"{slide_name}.rels"

    if rels_path.exists():
        rels = rels_path.read_text(encoding="utf-8")
    else:
        rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
                '</Relationships>')

    rids = [int(m) for m in re.findall(r'Id="rId(\d+)"', rels)]
    rid = f"rId{max(rids) + 1 if rids else 1}"
    rel = (f'<Relationship Id="{rid}" '
           'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
           f'Target="../charts/{chart_file}"/>')
    rels = rels.replace("</Relationships>", f"  {rel}\n</Relationships>")
    rels_path.write_text(rels, encoding="utf-8")
    return rid


def _insert_graphic_frame(unpacked, slide_name, rid, area_emu, chart_index):
    slide_path = unpacked / "ppt" / "slides" / slide_name
    xml = slide_path.read_text(encoding="utf-8")
    if "</p:spTree>" not in xml:
        raise SystemExit(f"Error: no <p:spTree> in {slide_name}")

    ids = [int(m) for m in re.findall(r'<p:cNvPr[^>]*\bid="(\d+)"', xml)]
    new_id = max(ids) + 1 if ids else 2
    x, y, w, h = area_emu

    frame = (
        '<p:graphicFrame>'
        '<p:nvGraphicFramePr>'
        f'<p:cNvPr id="{new_id}" name="Chart {chart_index}"/>'
        '<p:cNvGraphicFramePr><a:graphicFrameLocks noGrp="1"/></p:cNvGraphicFramePr>'
        '<p:nvPr/>'
        '</p:nvGraphicFramePr>'
        f'<p:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{w}" cy="{h}"/></p:xfrm>'
        '<a:graphic>'
        '<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
        '<c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        f'r:id="{rid}"/>'
        '</a:graphicData>'
        '</a:graphic>'
        '</p:graphicFrame>'
    )
    xml = xml.replace("</p:spTree>", f"{frame}</p:spTree>", 1)
    slide_path.write_text(xml, encoding="utf-8")


def _slide_size(unpacked):
    pres = (unpacked / "ppt" / "presentation.xml").read_text(encoding="utf-8")
    m = re.search(r'<p:sldSz[^>]*\bcx="(\d+)"[^>]*\bcy="(\d+)"', pres)
    if not m:
        m2 = re.search(r'<p:sldSz[^>]*\bcy="(\d+)"[^>]*\bcx="(\d+)"', pres)
        if m2:
            return int(m2.group(2)), int(m2.group(1))
        return 12192000, 6858000  # 16:9 default
    return int(m.group(1)), int(m.group(2))


def _take_over_existing_chart(unpacked, slide_name):
    """If the slide already holds a chart graphicFrame, capture its geometry,
    remove the frame, and drop its chart relationship (orphaning the old chart part
    for clean.py). Returns (x, y, w, h) in EMU, or None if there was no chart.

    This is what makes the "wrong plot type" path land in the right spot: copy the
    closest chart sample, then call this to swap its chart for a freshly-authored one
    of the needed type at the same position and size."""
    slide_path = unpacked / "ppt" / "slides" / slide_name
    dom = minidom.parse(str(slide_path))

    frame = None
    rid = None
    for gf in dom.getElementsByTagName("p:graphicFrame"):
        gd = gf.getElementsByTagName("a:graphicData")
        if gd and gd[0].getAttribute("uri") == CHART_URI:
            chart_ref = gf.getElementsByTagName("c:chart")
            if chart_ref:
                rid = chart_ref[0].getAttribute("r:id")
            frame = gf
            break
    if frame is None:
        return None

    geom = None
    xfrm = frame.getElementsByTagName("p:xfrm")
    if xfrm:
        off = xfrm[0].getElementsByTagName("a:off")
        ext = xfrm[0].getElementsByTagName("a:ext")
        if off and ext:
            geom = (int(off[0].getAttribute("x")), int(off[0].getAttribute("y")),
                    int(ext[0].getAttribute("cx")), int(ext[0].getAttribute("cy")))

    frame.parentNode.removeChild(frame)
    slide_path.write_bytes(dom.toxml(encoding="utf-8"))

    if rid:
        rels_path = unpacked / "ppt" / "slides" / "_rels" / f"{slide_name}.rels"
        if rels_path.exists():
            rels = rels_path.read_text(encoding="utf-8")
            rels = re.sub(rf'\s*<Relationship[^>]*\bId="{re.escape(rid)}"[^>]*/>', "", rels)
            rels_path.write_text(rels, encoding="utf-8")

    return geom


def _resolve_area(unpacked, area_in, inherited_geom=None):
    if area_in:
        try:
            x, y, w, h = (float(v) for v in area_in.split(","))
        except ValueError:
            raise SystemExit('Error: --area-in must be "x,y,w,h" in inches')
        return tuple(int(round(v * EMU_PER_INCH)) for v in (x, y, w, h))
    if inherited_geom:
        return inherited_geom
    sld_w, sld_h = _slide_size(unpacked)
    # Default: a content box below a title band.
    x = int(round(sld_w * 0.08))
    top = int(round(sld_h * 0.24))
    bottom = int(round(sld_h * 0.08))
    w = sld_w - 2 * x
    h = sld_h - top - bottom
    return x, top, w, h


def main():
    ap = argparse.ArgumentParser(
        description="Author a native chart from YAML data and embed it in a slide.")
    ap.add_argument("unpacked_dir")
    ap.add_argument("slide", help="slide file name, e.g. slide8.xml")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--yaml", help="slide-deck YAML; the data entry is extracted directly (no temp JSON)")
    src.add_argument("--data-json", help="path to one pre-extracted YAML data entry as JSON")
    ap.add_argument("--index", type=int, help="0-based slide index in --yaml")
    ap.add_argument("--data", help="data key within the slide (omit if it has a single data entry)")
    ap.add_argument("--area-in", default=None,
                    help='frame box "x,y,w,h" in inches; default = content area below a title')
    ap.add_argument("--title", default=None, help="optional chart title")
    args = ap.parse_args()

    unpacked = Path(args.unpacked_dir)
    if not (unpacked / "ppt" / "slides" / args.slide).exists():
        raise SystemExit(f"Error: {unpacked / 'ppt' / 'slides' / args.slide} not found")

    if args.yaml:
        if args.index is None:
            ap.error("--index is required with --yaml")
        from office.yaml_entry import load_data_entry
        entry = load_data_entry(args.yaml, args.index, args.data)
    else:
        entry = json.loads(Path(args.data_json).read_text(encoding="utf-8"))
    kind, cats, series = _normalize(entry)

    charts_dir = unpacked / "ppt" / "charts"
    emb_dir = unpacked / "ppt" / "embeddings"
    charts_dir.mkdir(exist_ok=True)
    (charts_dir / "_rels").mkdir(exist_ok=True)
    emb_dir.mkdir(exist_ok=True)

    n = _next_index(charts_dir, "chart", ".xml")
    chart_file = f"chart{n}.xml"
    xn = _next_index(emb_dir, "data_chart", ".xlsx")
    xlsx_file = f"data_chart{xn}.xlsx"

    # Embedded workbook + chart's own rels (chart -> workbook).
    write_workbook(emb_dir / xlsx_file, cats, series)
    chart_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
        f'  <Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/package" '
        f'Target="../embeddings/{xlsx_file}"/>\n'
        '</Relationships>'
    )
    (charts_dir / "_rels" / f"{chart_file}.rels").write_text(chart_rels, encoding="utf-8")

    # The chart part itself.
    chart_xml = build_chart_xml(kind, cats, series, args.title, "rId1")
    (charts_dir / chart_file).write_text(chart_xml, encoding="utf-8")

    # Register content types, link the slide to the chart, drop the frame in.
    # If a chart already lives on the slide (e.g. a copied wrong-type sample),
    # take over its slot: inherit its geometry and remove it.
    inherited_geom = _take_over_existing_chart(unpacked, args.slide)

    _ensure_content_types(unpacked, f"/ppt/charts/{chart_file}")
    rid = _add_slide_chart_rel(unpacked, args.slide, chart_file)
    area = _resolve_area(unpacked, args.area_in, inherited_geom)
    _insert_graphic_frame(unpacked, args.slide, rid, area, n)

    if args.area_in:
        placement = "at --area-in"
    elif inherited_geom:
        placement = "in place of the sample's chart (inherited position/size)"
    else:
        placement = "in a default content box"
    print(f"Built ppt/charts/{chart_file} ({kind}) and embedded it on {args.slide} "
          f"{placement}: {len(cats)} categories x {len(series)} series "
          f"(from scratch, theme-coloured).")


if __name__ == "__main__":
    main()
