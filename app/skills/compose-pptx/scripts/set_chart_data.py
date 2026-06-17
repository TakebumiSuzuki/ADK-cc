"""Give a slide its own chart and write YAML data into it (raw-XML, no python-pptx).

Why this exists: add_slide.py duplicates a slide by copying its XML *and rels*, so
a duplicated chart slide still points at the ORIGINAL chart part. Two slides built
from the same chart sample therefore share one chart part — editing one would
corrupt the other.

So this script first FORKS the chart part (chart{N}.xml + its embedded workbook get
fresh copies, the slide is repointed, Content_Types updated), then rewrites the
forked chart's cached data and regenerates the embedded .xlsx (via openpyxl, so
"Edit Data" in PowerPoint stays consistent). The sample chart's axes, colours and
styling are reused untouched — only the data changes.

It reuses the sample's plot type (bar/line/...). Route each YAML chart to a sample
of the matching type in the mapping step; if the YAML type and the sample's plot
type disagree this script warns but still fills the data.

Usage:
    # Preferred: read the data entry straight from the slide-deck YAML (no temp JSON).
    python set_chart_data.py <unpacked_dir> <slideN.xml> --yaml <deck.yaml> --index N [--data KEY]
    # Or pass a pre-extracted entry as JSON.
    python set_chart_data.py <unpacked_dir> <slideN.xml> --data-json <entry.json>

With `--yaml/--index`, the script extracts `slides[N].data[KEY]` itself (KEY may
be omitted when the slide has a single data entry), so values are never retyped
by hand. `entry.json` is one YAML `data:` entry, e.g.
    {"type":"bar_chart","categories":["Q1","Q2"],
     "series":[{"name":"2024","values":[10,20]},{"name":"2025","values":[12,25]}]}
    {"type":"histogram","bins":["0-20","20-40"],"frequencies":[3,12]}
"""

import argparse
import copy
import json
import re
import shutil
from pathlib import Path

import defusedxml.minidom as minidom
from openpyxl import Workbook

PLOT_TAGS = ("c:barChart", "c:lineChart", "c:areaChart", "c:pieChart",
             "c:scatterChart", "c:radarChart", "c:doughnutChart")
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"


# --------------------------------------------------------------------------- #
# Normalise the YAML entry into (categories, [(series_name, [values])])
# --------------------------------------------------------------------------- #
def _normalize(entry):
    etype = entry.get("type")
    if etype == "histogram":
        cats = [str(b) for b in entry.get("bins", [])]
        series = [(entry.get("y_axis", "Frequency"),
                   [float(v) for v in entry.get("frequencies", [])])]
    else:  # bar_chart / line_chart
        cats = [str(c) for c in entry.get("categories", [])]
        series = [(s.get("name", f"Series {i+1}"),
                   [float(v) for v in s.get("values", [])])
                  for i, s in enumerate(entry.get("series", []))]
    if not series:
        raise SystemExit("Error: chart entry has no series/frequencies")
    return cats, series


# --------------------------------------------------------------------------- #
# Fork the chart part referenced by the slide
# --------------------------------------------------------------------------- #
def _slide_chart_rel(unpacked, slide_name):
    rels = unpacked / "ppt" / "slides" / "_rels" / f"{slide_name}.rels"
    dom = minidom.parse(str(rels))
    for rel in dom.getElementsByTagName("Relationship"):
        if rel.getAttribute("Type").endswith("/chart"):
            return rels, dom, rel
    return rels, dom, None


def _next_index(charts_dir, stem, ext):
    nums = [int(m.group(1)) for f in charts_dir.glob(f"{stem}*{ext}")
            if (m := re.match(rf"{stem}(\d+){re.escape(ext)}", f.name))]
    return max(nums, default=0) + 1


def fork_chart(unpacked, slide_name):
    """Copy the slide's chart part (+ embedded workbook) to fresh names, repoint
    the slide, register Content_Types. Returns the new chart Path."""
    rels_path, rels_dom, rel = _slide_chart_rel(unpacked, slide_name)
    if rel is None:
        raise SystemExit(f"Error: {slide_name} references no chart to fill")

    charts_dir = unpacked / "ppt" / "charts"
    old_chart = (charts_dir / Path(rel.getAttribute("Target")).name).resolve()
    n = _next_index(charts_dir, "chart", ".xml")
    new_chart = charts_dir / f"chart{n}.xml"
    shutil.copy2(old_chart, new_chart)

    # Fork the chart's own rels, repointing its embedded workbook to a fresh copy.
    old_crels = charts_dir / "_rels" / f"{old_chart.name}.rels"
    new_crels = charts_dir / "_rels" / f"chart{n}.xml.rels"
    if old_crels.exists():
        cdom = minidom.parse(str(old_crels))
        for r in cdom.getElementsByTagName("Relationship"):
            if r.getAttribute("Type").endswith("/package"):  # embedded xlsx
                # .rels Targets resolve relative to the PART's dir (ppt/charts), not _rels.
                old_xlsx = (charts_dir / r.getAttribute("Target")).resolve()
                emb_dir = unpacked / "ppt" / "embeddings"
                xn = _next_index(emb_dir, "data_chart", ".xlsx")
                new_xlsx = emb_dir / f"data_chart{xn}.xlsx"
                if old_xlsx.exists():
                    shutil.copy2(old_xlsx, new_xlsx)
                rel_target = f"../embeddings/{new_xlsx.name}"
                r.setAttribute("Target", rel_target)
        new_crels.write_bytes(cdom.toxml(encoding="utf-8"))

    # Repoint the slide at the forked chart, and register it in Content_Types.
    rel.setAttribute("Target", f"../charts/chart{n}.xml")
    rels_path.write_bytes(rels_dom.toxml(encoding="utf-8"))

    ct_path = unpacked / "[Content_Types].xml"
    ct = ct_path.read_text(encoding="utf-8")
    part = f"/ppt/charts/chart{n}.xml"
    if part not in ct:
        ov = (f'<Override PartName="{part}" '
              f'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
        ct = ct.replace("</Types>", f"  {ov}\n</Types>")
        ct_path.write_text(ct, encoding="utf-8")

    return new_chart


# --------------------------------------------------------------------------- #
# Rewrite the chart XML caches
# --------------------------------------------------------------------------- #
def _set_text(parent, tag, value):
    """Replace parent's single <tag> child text content."""
    for c in parent.getElementsByTagName(tag):
        while c.firstChild:
            c.removeChild(c.firstChild)
        c.appendChild(parent.ownerDocument.createTextNode(str(value)))
        return


def _rebuild_cache(dom, ref_el, cache_tag, values, formula, number=False):
    """Rewrite a <c:strRef>/<c:numRef> child: set <c:f> and the str/num cache pts."""
    # formula
    f = ref_el.getElementsByTagName("c:f")
    if f:
        _set_text(ref_el, "c:f", formula)
    # cache
    caches = ref_el.getElementsByTagName(cache_tag)
    if not caches:
        return
    cache = caches[0]
    # wipe existing pt + ptCount, keep formatCode if present
    for node in list(cache.childNodes):
        if node.nodeType == node.ELEMENT_NODE and node.tagName in ("c:pt", "c:ptCount"):
            cache.removeChild(node)
    ptcount = dom.createElementNS(C_NS, "c:ptCount")
    ptcount.setAttribute("val", str(len(values)))
    cache.insertBefore(ptcount, cache.firstChild)
    for i, v in enumerate(values):
        pt = dom.createElementNS(C_NS, "c:pt")
        pt.setAttribute("idx", str(i))
        vv = dom.createElementNS(C_NS, "c:v")
        vv.appendChild(dom.createTextNode(_num(v) if number else str(v)))
        pt.appendChild(vv)
        cache.appendChild(pt)


def _num(v):
    f = float(v)
    return str(int(f)) if f.is_integer() else repr(f)


def _col_letter(i):  # 0->B, 1->C ... (data columns start at B; A is categories)
    return chr(ord("B") + i)


def write_chart_data(chart_path, cats, series):
    dom = minidom.parse(str(chart_path))
    plot = None
    for tag in PLOT_TAGS:
        els = dom.getElementsByTagName(tag)
        if els:
            plot = els[0]
            break
    if plot is None:
        raise SystemExit("Error: no recognised plot element in chart")

    sers = [c for c in plot.childNodes
            if c.nodeType == c.ELEMENT_NODE and c.tagName == "c:ser"]
    if not sers:
        raise SystemExit("Error: chart has no <c:ser> to use as a template")

    # Grow/shrink the series list to match the data (clone the last ser as needed).
    while len(sers) < len(series):
        clone = sers[-1].cloneNode(deep=True)
        plot.insertBefore(clone, sers[-1].nextSibling)
        sers.append(clone)
    for extra in sers[len(series):]:
        plot.removeChild(extra)
    sers = sers[:len(series)]

    n = len(cats)
    for si, (sname, vals) in enumerate(series):
        ser = sers[si]
        # idx / order
        for tag in ("c:idx", "c:order"):
            for e in ser.getElementsByTagName(tag):
                e.setAttribute("val", str(si)); break
        col = _col_letter(si)
        # series name (tx > strRef)
        tx = ser.getElementsByTagName("c:tx")
        if tx:
            ref = tx[0].getElementsByTagName("c:strRef")
            if ref:
                _rebuild_cache(dom, ref[0], "c:strCache", [sname], f"Sheet1!${col}$1")
        # categories (cat > strRef|numRef)
        cat = ser.getElementsByTagName("c:cat")
        if cat:
            sref = cat[0].getElementsByTagName("c:strRef")
            nref = cat[0].getElementsByTagName("c:numRef")
            rng = f"Sheet1!$A$2:$A${n+1}"
            if sref:
                _rebuild_cache(dom, sref[0], "c:strCache", cats, rng)
            elif nref:
                _rebuild_cache(dom, nref[0], "c:numCache", cats, rng, number=True)
        # values (val > numRef)
        val = ser.getElementsByTagName("c:val")
        if val:
            nref = val[0].getElementsByTagName("c:numRef")
            if nref:
                _rebuild_cache(dom, nref[0], "c:numCache", vals,
                               f"Sheet1!${col}$2:${col}${n+1}", number=True)

    chart_path.write_bytes(dom.toxml(encoding="utf-8"))


def write_workbook(unpacked, chart_path, cats, series):
    """Regenerate the chart's embedded .xlsx so Edit-Data matches the caches."""
    crels = chart_path.parent / "_rels" / f"{chart_path.name}.rels"
    if not crels.exists():
        return
    dom = minidom.parse(str(crels))
    target = None
    for r in dom.getElementsByTagName("Relationship"):
        if r.getAttribute("Type").endswith("/package"):
            # Targets are relative to the part dir (ppt/charts), i.e. _rels' parent.
            target = (crels.parent.parent / r.getAttribute("Target")).resolve()
    if target is None:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=None)
    for si, (sname, _vals) in enumerate(series):
        ws.cell(row=1, column=2 + si, value=sname)
    for ri, cat in enumerate(cats):
        ws.cell(row=2 + ri, column=1, value=cat)
        for si, (_sname, vals) in enumerate(series):
            ws.cell(row=2 + ri, column=2 + si,
                    value=(vals[ri] if ri < len(vals) else None))
    wb.save(str(target))


def main():
    ap = argparse.ArgumentParser(description="Fork a slide's chart and fill it with YAML data.")
    ap.add_argument("unpacked_dir")
    ap.add_argument("slide", help="slide file name, e.g. slide8.xml")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--yaml", help="slide-deck YAML; the data entry is extracted directly (no temp JSON)")
    src.add_argument("--data-json", help="path to one pre-extracted YAML data entry as JSON")
    ap.add_argument("--index", type=int, help="0-based slide index in --yaml")
    ap.add_argument("--data", help="data key within the slide (omit if it has a single data entry)")
    args = ap.parse_args()

    unpacked = Path(args.unpacked_dir)
    if args.yaml:
        if args.index is None:
            ap.error("--index is required with --yaml")
        from office.yaml_entry import load_data_entry
        entry = load_data_entry(args.yaml, args.index, args.data)
    else:
        entry = json.loads(Path(args.data_json).read_text(encoding="utf-8"))
    cats, series = _normalize(entry)

    new_chart = fork_chart(unpacked, args.slide)
    write_chart_data(new_chart, cats, series)
    write_workbook(unpacked, new_chart, cats, series)

    print(f"Filled {new_chart.relative_to(unpacked)} on {args.slide}: "
          f"{len(cats)} categories x {len(series)} series (independent fork).")


if __name__ == "__main__":
    main()
